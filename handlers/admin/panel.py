from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ContextTypes
import logging
from config import ADMIN_IDS, SUPER_ADMIN_IDS, MODERATOR_IDS
from utils.admin_keyboards import get_admin_panel_keyboard, get_moderator_panel_keyboard, get_admin_panel_text, get_moderator_panel_text

logger = logging.getLogger(__name__)

async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Главная панель администратора"""
    try:
        user_id = update.effective_user.id
        # ВРЕМЕННОЕ ЛОГИРОВАНИЕ ДЛЯ ОТЛАДКИ
        # Проверяем права доступа
        if user_id not in ADMIN_IDS:
            await update.message.reply_text("У вас нет прав администратора.")
            return
        
        # Определяем уровень доступа
        if user_id in SUPER_ADMIN_IDS:
            # Полная админ панель
            reply_markup = get_admin_panel_keyboard()
            text = get_admin_panel_text()
        elif user_id in MODERATOR_IDS:
            # Только модерация
            reply_markup = get_moderator_panel_keyboard()
            text = get_moderator_panel_text()
        else:
            await update.message.reply_text("У вас нет прав администратора.")
            return
        
        await update.message.reply_text(text, reply_markup=reply_markup)
        
    except Exception as e:
        logger.error(f"Error in admin_panel: {e}")
        await update.message.reply_text("Произошла ошибка.")

def is_admin(user_id: int) -> bool:
    """Проверка прав администратора (любой уровень)"""
    return user_id in ADMIN_IDS

def is_super_admin(user_id: int) -> bool:
    """Проверка прав главного администратора"""
    return user_id in SUPER_ADMIN_IDS

def is_moderator(user_id: int) -> bool:
    """Проверка прав модератора"""
    return user_id in MODERATOR_IDS
    
async def export_all_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт всех пользователей в Excel - ТОЛЬКО ДЛЯ ГЛАВНОГО АДМИНА"""
    try:
        query = update.callback_query
        await query.answer()
        
        user_id = query.from_user.id
        
        # Проверяем, что это главный админ
        if not is_super_admin(user_id):
            await query.edit_message_text("Нет прав доступа. Эта функция доступна только главному администратору.")
            return
        
        # Получаем всех пользователей
        from database.connection import db
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT telegram_id, full_name, phone_number, 
                       player_level, created_at
                FROM users 
                WHERE telegram_id > 0
                ORDER BY created_at DESC
            """)
            
            users = cursor.fetchall()
        
        if not users:
            await query.edit_message_text("Нет пользователей для экспорта")
            return
        
        # Создаем Excel файл
        import io
        from datetime import datetime
        import xlsxwriter
        from levels import get_level_name, get_category_by_level
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Пользователи')
        
        # Форматы
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'align': 'center'
        })
        
        cell_format = workbook.add_format({'bg_color': '#F8F9FA'})
        
        # Заголовки (ОБНОВЛЕНО - убрали "Возрастная категория", добавили "Уровень игры")
        headers = ['Telegram ID', 'ФИО', 'Телефон', 'Уровень игры', 'Категория', 'Дата регистрации']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Данные пользователей
        for row, user in enumerate(users, 1):
            telegram_id = user[0]
            full_name = user[1]
            phone = user[2]
            player_level = user[3]
            created_at = user[4]
            
            # Форматируем дату регистрации
            try:
                created_date = datetime.fromisoformat(created_at)
                formatted_date = created_date.strftime('%d.%m.%Y')
            except:
                formatted_date = created_at[:10] if created_at else ''
            
            # Получаем название уровня и категорию
            if player_level:
                level_display = player_level  # ← ТЕПЕРЬ ПРОСТО "1.5"
                category = get_category_by_level(player_level)
                category_display = f"Категория {category}" if category else ""
            else:
                level_display = ""  # ← Пустая строка вместо "Не установлен"
                category_display = ""
            
            worksheet.write(row, 0, telegram_id, cell_format)
            worksheet.write(row, 1, full_name, cell_format)
            worksheet.write(row, 2, phone, cell_format)
            worksheet.write(row, 3, level_display, cell_format)
            worksheet.write(row, 4, category_display, cell_format)
            worksheet.write(row, 5, formatted_date, cell_format)
        
        # Автоподбор ширины колонок
        worksheet.set_column('A:A', 12)  # Telegram ID
        worksheet.set_column('B:B', 25)  # ФИО
        worksheet.set_column('C:C', 15)  # Телефон
        worksheet.set_column('D:D', 20)  # Уровень игры
        worksheet.set_column('E:E', 15)  # Категория
        worksheet.set_column('F:F', 15)  # Дата регистрации
        
        workbook.close()
        output.seek(0)
        
        # Отправляем файл
        filename = f"all_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=output,
            filename=filename,
            caption=f"📊 Все пользователи бота\nВсего пользователей: {len(users)}"
        )
        
        # Отправляем админ панель отдельным сообщением
        from utils.admin_keyboards import get_admin_panel_keyboard, get_admin_panel_text
        
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=get_admin_panel_text(),
            reply_markup=get_admin_panel_keyboard()
        )
        
        # Удаляем исходное сообщение
        await query.delete_message()
        
    except Exception as e:
        logger.error(f"Error in export_all_users: {e}")
        try:
            await query.edit_message_text("Произошла ошибка при экспорте")
        except:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Произошла ошибка при экспорте"
            )
            
async def import_users_levels(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Импорт уровней пользователей из Excel - ТОЛЬКО ДЛЯ ГЛАВНОГО АДМИНА"""
    try:
        query = update.callback_query
        await query.answer()
        
        user_id = query.from_user.id
        
        if not is_super_admin(user_id):
            await query.edit_message_text("Нет прав доступа. Эта функция доступна только главному администратору.")
            return
        
        keyboard = [
            [InlineKeyboardButton("❌ Отмена", callback_data="admin_panel_return")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "📤 Импорт уровней пользователей\n\n"
            "Отправьте Excel файл (.xlsx) с колонками:\n"
            "• Telegram ID\n"
            "• ФИО\n"
            "• Телефон\n"
            "• Уровень игры (например: 1.5, 2.25, 3.0)\n\n"
            "⚠️ Пустые уровни будут сброшены в NULL",
            reply_markup=reply_markup
        )
        
        # Сохраняем состояние ожидания файла
        context.user_data['awaiting_import'] = True
        
    except Exception as e:
        logger.error(f"Error in import_users_levels: {e}")
        await query.edit_message_text("Произошла ошибка")


async def handle_import_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка загруженного Excel файла для импорта уровней"""
    try:
        # Проверяем что ждём файл
        if not context.user_data.get('awaiting_import'):
            return
        
        user_id = update.effective_user.id
        
        if not is_super_admin(user_id):
            await update.message.reply_text("Нет прав доступа.")
            return
        
        # Проверяем что это документ
        if not update.message.document:
            await update.message.reply_text("Пожалуйста, отправьте Excel файл (.xlsx)")
            return
        
        # Проверяем расширение
        filename = update.message.document.file_name
        if not filename.endswith('.xlsx'):
            await update.message.reply_text("Файл должен быть в формате .xlsx")
            return
        
        await update.message.reply_text("⏳ Обрабатываю файл...")
        
        # Скачиваем файл
        file = await context.bot.get_file(update.message.document.file_id)
        import io
        file_bytes = io.BytesIO()
        await file.download_to_memory(file_bytes)
        file_bytes.seek(0)
        
        # Читаем Excel
        import openpyxl
        workbook = openpyxl.load_workbook(file_bytes)
        sheet = workbook.active
        
        # Проверяем заголовки
        headers = [cell.value for cell in sheet[1]]
        
        required_headers = ['Telegram ID', 'ФИО', 'Телефон', 'Уровень игры']
        if not all(h in headers for h in required_headers):
            await update.message.reply_text(
                f"❌ Ошибка: файл должен содержать колонки:\n"
                f"• Telegram ID\n"
                f"• ФИО\n"
                f"• Телефон\n"
                f"• Уровень игры"
            )
            context.user_data['awaiting_import'] = False
            return
        
        # Находим индексы колонок
        telegram_id_col = headers.index('Telegram ID')
        level_col = headers.index('Уровень игры')
        
        # Обрабатываем строки
        from database.connection import db
        from datetime import datetime
        
        updated_count = 0
        skipped_count = 0
        error_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                telegram_id = row[telegram_id_col]
                new_level = row[level_col]
                
                if not telegram_id:
                    skipped_count += 1
                    continue
                
                # Преобразуем telegram_id в int
                telegram_id = int(telegram_id)
                
                # Обрабатываем уровень
                if new_level:
                    # Убираем возможные пробелы и проверяем формат
                    new_level = str(new_level).strip()
                    if new_level.lower() in ['', 'не установлен', 'none', 'null']:
                        new_level = None
                else:
                    new_level = None
                
                # Обновляем в БД
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    if new_level:
                        cursor.execute("""
                            UPDATE users 
                            SET player_level = ?,
                                player_level_updated_at = ?,
                                player_level_updated_by = ?
                            WHERE telegram_id = ?
                        """, (new_level, datetime.now(), user_id, telegram_id))
                    else:
                        # Сбрасываем уровень
                        cursor.execute("""
                            UPDATE users 
                            SET player_level = NULL,
                                player_level_updated_at = ?,
                                player_level_updated_by = ?
                            WHERE telegram_id = ?
                        """, (datetime.now(), user_id, telegram_id))
                    
                    conn.commit()
                    
                    if cursor.rowcount > 0:
                        updated_count += 1
                    else:
                        skipped_count += 1
                        logger.warning(f"User {telegram_id} not found in database")
                
            except Exception as e:
                error_count += 1
                logger.error(f"Error processing row {row}: {e}")
        
        # Отчёт
        from utils.admin_keyboards import get_admin_panel_keyboard, get_admin_panel_text
        
        report = (
            f"✅ Импорт завершён!\n\n"
            f"📊 Статистика:\n"
            f"• Обновлено: {updated_count}\n"
            f"• Пропущено: {skipped_count}\n"
            f"• Ошибок: {error_count}\n\n"
            f"Файл: {filename}"
        )
        
        await update.message.reply_text(report)
        await update.message.reply_text(
            get_admin_panel_text(),
            reply_markup=get_admin_panel_keyboard()
        )
        
        context.user_data['awaiting_import'] = False
        
    except Exception as e:
        logger.error(f"Error in handle_import_file: {e}")
        await update.message.reply_text(f"❌ Ошибка при обработке файла: {e}")
        context.user_data['awaiting_import'] = False